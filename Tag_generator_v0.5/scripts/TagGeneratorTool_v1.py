#!/usr/bin/env python3
"""Overview: read PLC variable names, recognize their meaning, and generate tag rows."""

import sys
import csv
from pathlib import Path
from openpyxl import load_workbook

from recognition_v1 import (
    load_yaml,
    build_dictionary,
    RecognitionEngine,
    result_to_csv_row,
    get_recognition_statistics
)

from naming_v1 import (
    NamingEngine
)

PROJECT_DIR = Path(__file__).resolve().parent.parent
DICTIONARY_FILE = PROJECT_DIR / 'yaml' / 'Master_Dictionary_v2.1.yaml'
RECOGNITION_FILE = PROJECT_DIR / 'yaml' / 'Recognition_Engine_v0.3.yaml'
TEMPLATE_FILE = PROJECT_DIR / 'results' / 'templates' / 'Tag_Generator_Template_v1.xlsx'

def load_config():
    """Load all configuration files."""

    dictionary = build_dictionary(
        load_yaml(DICTIONARY_FILE)
    )

    recognitionFile = load_yaml(
        RECOGNITION_FILE
    )

    return dictionary, recognitionFile

def parse_arguments():

    if len(sys.argv) != 3:
        print(
            "ERROR: USAGE: py generatorTool.py input.txt output.xlsx"
        )
        sys.exit(1)

    input_file = PROJECT_DIR / 'inputs' / sys.argv[1]
    output_file = PROJECT_DIR / 'results' / 'generated' / sys.argv[2]

    return input_file, output_file

def recognize_variables(input_file, recognition_engine):
    """Process all variables from the input file."""

    recognized_variables = []

    with open(
        input_file,
        "r",
        encoding="utf-8"
    ) as f:

        for line in f:

            variable = line.strip()

            if not variable:
                continue

            recognized_variables.append(
                recognition_engine.process_variable(variable)
            )

    return recognized_variables

def name_variables(recognized_variables, naming_engine):
    """Convert recognized variables to named tags using the naming engine."""
    named_variables = []
    for variable in recognized_variables:
        named_variables.append(naming_engine.build(variable))

    return named_variables

def export_excel(
        generated_tags,
        template_file,
        output_file):

    generate_statistics = True

    csv_rows = []

    for tag in generated_tags:

        csv_rows.append(result_to_csv_row(tag))


    wb = load_workbook(template_file)
    ws = wb.active

    # Leer cabeceras de la fila 11
    headers = {
        cell.value: cell.column
        for cell in ws[11]
        if cell.value
    }

    for row_num, row_data in enumerate(csv_rows, start=12):

        for field, value in row_data.items():

            if field not in headers:
                continue

            ws.cell(
                row=row_num,
                column=headers[field],
                value=value
            )

    wb.save(output_file)

    if generate_statistics:
        get_recognition_statistics(
            recognized_variables
        )

if __name__ == "__main__":

    input_file, output_file = parse_arguments()

    dictionary, recognitionFile = load_config()

    recognition_engine = RecognitionEngine(
        dictionary,
        recognitionFile
    )

    naming_engine = NamingEngine(
        dictionary
    )

    recognized_variables = recognize_variables(input_file, recognition_engine)

    generated_variables = name_variables(recognized_variables, naming_engine)

    export_excel(generated_variables, TEMPLATE_FILE, output_file)
